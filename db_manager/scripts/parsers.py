import hashlib
import json
import logging
import os
from collections.abc import Iterable, Iterator
from dataclasses import dataclass
from datetime import date, datetime

from langchain_core.documents import Document

logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class ParsedDoc:
    document: Document
    chunk_kind: str  # table | formulas | text | ocr


def sha256_file(path: str, chunk_size: int = 1024 * 1024) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        while chunk := f.read(chunk_size):
            h.update(chunk)
    return h.hexdigest()


def _stringify_cell(v) -> str:
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    if isinstance(v, float):
        return f"{v:.10g}"
    return str(v)


def _load_email_meta(path: str) -> dict:
    meta_path = path + ".meta.json"
    if not os.path.exists(meta_path):
        return {}
    try:
        with open(meta_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as exc:
        logger.warning("Не удалось прочитать метаданные %s: %s", meta_path, exc)
        return {}


def parse_xlsx(
    path: str,
    *,
    rows_per_chunk: int = 50,
    max_cols: int = 60,
) -> Iterator[ParsedDoc]:
    """
    Парсит XLSX в два вида чанков:
    - table: строки данных с повторяющейся шапкой
    - formulas: ячейки с формулами и их вычисленными значениями
    """
    from openpyxl import load_workbook

    file_name = os.path.basename(path)
    file_hash = sha256_file(path)
    email_meta = _load_email_meta(path)

    base_meta = {
        "source": file_name,
        "path": path,
        "file_type": "xlsx",
        "sha256": file_hash,
        **email_meta,
    }

    wb_formulas = load_workbook(path, data_only=False, read_only=True)
    wb_values = load_workbook(path, data_only=True, read_only=True)

    for ws in wb_formulas.worksheets:
        ws_v = wb_values[ws.title]
        max_row = min(ws.max_row or 0, 20_000)
        max_col = min(ws.max_column or 0, max_cols)
        if max_row <= 0 or max_col <= 0:
            continue

        sheet_meta = {**base_meta, "sheet": ws.title}

        # --- Формулы ---
        formula_lines: list[str] = []
        for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cached = _stringify_cell(ws_v[cell.coordinate].value)
                    formula_lines.append(f"{cell.coordinate}\t{cell.value}\tvalue={cached}")

        if formula_lines:
            content = "\n".join([
                f"Файл: {file_name}", f"Лист: {ws.title}", "Тип: formulas", "",
                *formula_lines[:20_000],
            ])
            yield ParsedDoc(
                document=Document(
                    page_content=content,
                    metadata={**sheet_meta, "chunk_kind": "formulas"},
                ),
                chunk_kind="formulas",
            )

        # --- Табличные данные ---
        header_row = next(
            ws.iter_rows(min_row=1, max_row=1, max_col=max_col, values_only=False),
            [],
        )
        header_line = "\t".join(_stringify_cell(c.value) for c in header_row).rstrip()

        buf: list[str] = []
        buf_start: int | None = None
        buf_end: int | None = None

        def _flush_table_chunk(row_start: int, row_end: int, lines: list[str]) -> ParsedDoc:
            content = "\n".join([
                f"Файл: {file_name}", f"Лист: {ws.title}", "Тип: table", "",
                "HEADER\t" + (header_line or ""),
                *("ROW\t" + l for l in lines),
            ])
            return ParsedDoc(
                document=Document(
                    page_content=content,
                    metadata={**sheet_meta, "chunk_kind": "table",
                               "row_start": row_start, "row_end": row_end},
                ),
                chunk_kind="table",
            )

        for row_idx, row in enumerate(
            ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col, values_only=False),
            start=2,
        ):
            line = "\t".join(_stringify_cell(c.value) for c in row).rstrip()
            if not line:
                continue
            if buf_start is None:
                buf_start = row_idx
            buf_end = row_idx
            buf.append(line)

            if len(buf) >= rows_per_chunk:
                yield _flush_table_chunk(buf_start, buf_end, buf)
                buf, buf_start, buf_end = [], None, None

        if buf:
            yield _flush_table_chunk(buf_start, buf_end, buf)


def parse_files(
    paths: Iterable[str],
    *,
    xlsx_rows_per_chunk: int = 50,
) -> Iterator[ParsedDoc]:
    """
    Единая точка входа для парсеров.
    - XLSX: встроенный парсер (таблицы + формулы)
    - PDF, DOCX, TXT: обрабатываются через DoclingLoader в init_db
    """
    for p in paths:
        if os.path.splitext(p)[1].lower() == ".xlsx":
            yield from parse_xlsx(p, rows_per_chunk=xlsx_rows_per_chunk)
